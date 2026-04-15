"""
DC_energy_modelling_v2.py
=========================
Data Center Power Consumption Model for VHP Integration
Integrated with Open-Meteo / ERA5 real weather data.

Usage
-----
Edit the CONFIG block below, then run:
    python3 DC_energy_modelling_v2.py

Weather data is fetched automatically from the Open-Meteo API (no API key needed).
Historical years use real ERA5 reanalysis data; the current year uses a hybrid of
real + statistical data; future years use Gaussian climatological sampling.
"""

# ─────────────────────────────────────────────────────────────────────────────
# USER CONFIGURATION  ← edit everything in this block
# ─────────────────────────────────────────────────────────────────────────────

CONFIG = {
    # ── Location ──────────────────────────────────────────────────────────────
    # Any city name accepted by Open-Meteo geocoding, e.g.:
    #   "Edinburgh", "London", "Stockholm", "Singapore", "New York", "Sydney"
    "city": "Edinburgh",

    # ── Year ──────────────────────────────────────────────────────────────────
    # Past years  → real ERA5 data
    # Current year → hybrid real + statistical
    # Future years → Gaussian climatological prediction
    "year": 2024,

    # ── Timestep ──────────────────────────────────────────────────────────────
    # 1   = hourly   (8,760  rows/year)
    # 0.5 = half-hourly (17,520 rows/year, interpolated from ERA5 hourly)
    "timestep_hours": 1,

    # ── Data Centre Parameters ────────────────────────────────────────────────
    "it_capacity_kw":    500,    # Peak IT capacity (kW). 1 MW = 1000, 50 MW = 50000
    "pue":               1.4,    # Power Usage Effectiveness (1.2 modern → 2.0 old)
    "base_utilization":  0.65,   # Baseline IT load as fraction of capacity (0-1)

    # ── Output ────────────────────────────────────────────────────────────────
    "output_csv":   True,        # Save hourly data to CSV
    "output_excel": True,        # Save to Excel (.xlsx) with Metadata + Data sheets
    "output_plots": True,        # Save PNG plots
    "output_dir":   ".",         # Directory for output files
}

# ─────────────────────────────────────────────────────────────────────────────
# END OF USER CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────


import time
import warnings
from datetime import date, timedelta
from typing import Optional

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")          # headless — no display needed
import matplotlib.pyplot as plt
import requests

warnings.filterwarnings("ignore")

# ── Open-Meteo endpoints ──────────────────────────────────────────────────────
_GEOCODING_URL   = "https://geocoding-api.open-meteo.com/v1/search"
_HISTORICAL_URL  = "https://archive-api.open-meteo.com/v1/archive"
_WEATHER_VARS    = ["temperature_2m"]      # only temperature needed
_CURRENT_YEAR    = date.today().year
_CURRENT_DATE    = date.today()


# ═════════════════════════════════════════════════════════════════════════════
# WEATHER FETCHING  (derived from weather_fetcher.py)
# ═════════════════════════════════════════════════════════════════════════════

def geocode_city(city_name: str) -> dict:
    """Resolve a city name to lat/lon/timezone via Open-Meteo Geocoding API."""
    print(f"\n[Weather] Geocoding '{city_name}' ...")
    resp = requests.get(
        _GEOCODING_URL,
        params={"name": city_name, "count": 5, "language": "en", "format": "json"},
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json()
    if not data.get("results"):
        raise ValueError(
            f"City '{city_name}' not found. "
            "Try a more specific name, e.g. 'Paris, France'."
        )
    r = data["results"][0]
    info = {
        "name":      r["name"],
        "country":   r.get("country", ""),
        "latitude":  r["latitude"],
        "longitude": r["longitude"],
        "timezone":  r.get("timezone", "UTC"),
        "elevation": r.get("elevation", 0),
    }
    print(
        f"         Found: {info['name']}, {info['country']} "
        f"(lat={info['latitude']:.4f}, lon={info['longitude']:.4f})"
    )
    return info


def _fetch_chunk(lat, lon, start, end, tz, retries=3) -> pd.Series:
    """Fetch temperature_2m for a date range; return a Series indexed by datetime."""
    params = {
        "latitude":          lat,
        "longitude":         lon,
        "start_date":        start,
        "end_date":          end,
        "hourly":            "temperature_2m",
        "timezone":          tz,
        "temperature_unit":  "celsius",
    }
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(_HISTORICAL_URL, params=params, timeout=60)
            resp.raise_for_status()
            break
        except requests.RequestException as exc:
            if attempt == retries:
                raise
            wait = 2 ** attempt
            print(f"         Retry {attempt}/{retries} after {wait}s ({exc})")
            time.sleep(wait)

    raw = resp.json()
    if "error" in raw:
        raise RuntimeError(f"API error: {raw.get('reason', raw)}")

    hourly = raw.get("hourly", {})
    if not hourly or "time" not in hourly:
        return pd.Series(dtype=float)

    series = pd.Series(
        hourly["temperature_2m"],
        index=pd.to_datetime(hourly["time"]),
        name="temperature_2m",
    )
    return series


def _fetch_year(lat, lon, year, tz) -> pd.Series:
    """Fetch a full calendar year of hourly temperature in two 6-month chunks."""
    chunks = [
        (f"{year}-01-01", f"{year}-06-30"),
        (f"{year}-07-01", f"{year}-12-31"),
    ]
    parts = []
    for start, end in chunks:
        s = _fetch_chunk(lat, lon, start, end, tz)
        if not s.empty:
            parts.append(s)
        time.sleep(0.3)
    if not parts:
        raise RuntimeError(f"No temperature data returned for {year}.")
    return pd.concat(parts).sort_index()


def _build_normals(lat, lon, tz, n_years=10, ref_end=None) -> pd.DataFrame:
    """Compute hourly climatological normals (mean ± std) over n_years."""
    if ref_end is None:
        ref_end = _CURRENT_YEAR - 1
    ref_start = ref_end - n_years + 1
    print(f"[Weather] Building climatological normals ({ref_start}-{ref_end}) ...")

    frames = []
    for yr in range(ref_start, ref_end + 1):
        print(f"          Fetching reference year {yr} ...", end="", flush=True)
        try:
            s = _fetch_year(lat, lon, yr, tz)
            df = s.to_frame()
            df["month"] = df.index.month
            df["day"]   = df.index.day
            df["hour"]  = df.index.hour
            frames.append(df)
            print(" ok")
        except Exception as exc:
            print(f" skipped ({exc})")

    if not frames:
        raise RuntimeError("Could not fetch any reference years for normals.")

    combined = pd.concat(frames)
    normals = (
        combined.groupby(["month", "day", "hour"])["temperature_2m"]
        .agg(["mean", "std"])
        .reset_index()
        .fillna(0)
    )
    return normals


def _sample_synthetic_year(normals: pd.DataFrame, year: int, seed=42) -> pd.Series:
    """Sample a synthetic temperature year from climatological normals."""
    rng = np.random.default_rng(seed)
    idx = pd.date_range(f"{year}-01-01", f"{year}-12-31 23:00:00", freq="h")
    df = pd.DataFrame({"month": idx.month, "day": idx.day, "hour": idx.hour}, index=idx)
    df = df.merge(normals, on=["month", "day", "hour"], how="left")
    df.index = idx
    noise  = rng.normal(0, df["std"].fillna(0).values)
    temps  = df["mean"].values + noise
    return pd.Series(temps, index=idx, name="temperature_2m")


def fetch_temperature(city: str, year: int) -> tuple[pd.Series, str, dict]:
    """
    Main entry point: fetch or predict hourly outdoor temperature for a city/year.

    Returns
    -------
    series   : pd.Series  - hourly temperature indexed by datetime
    mode     : str        - human-readable description of data source
    city_info: dict       - geocoding metadata
    """
    city_info = geocode_city(city)
    lat, lon, tz = city_info["latitude"], city_info["longitude"], city_info["timezone"]

    if year < 1940:
        raise ValueError("ERA5 data is only available from 1940 onwards.")

    if year < _CURRENT_YEAR:
        print(f"[Weather] Fetching historical ERA5 temperature for {year} ...")
        series = _fetch_year(lat, lon, year, tz)
        mode = f"Historical ERA5 (Open-Meteo) — {year}"

    elif year == _CURRENT_YEAR:
        yesterday    = _CURRENT_DATE - timedelta(days=1)
        actual_start = f"{year}-01-01"
        actual_end   = yesterday.strftime("%Y-%m-%d")
        print(f"[Weather] Fetching actual data {actual_start} → {actual_end} ...")
        actual = _fetch_chunk(lat, lon, actual_start, actual_end, tz)

        normals = _build_normals(lat, lon, tz, n_years=10, ref_end=year - 1)
        full_pred = _sample_synthetic_year(normals, year)
        predicted = full_pred[full_pred.index.date >= _CURRENT_DATE]

        series = pd.concat([actual, predicted]).sort_index()
        mode = (
            f"Hybrid: ERA5 Jan 1-{yesterday.strftime('%b %d')}, "
            f"predicted {_CURRENT_DATE.strftime('%b %d')}-Dec 31"
        )

    else:
        print(f"[Weather] Year {year} is in the future — generating prediction ...")
        normals = _build_normals(lat, lon, tz, n_years=10, ref_end=_CURRENT_YEAR - 1)
        series  = _sample_synthetic_year(normals, year)
        mode = (
            f"Predicted (climatological normals "
            f"{_CURRENT_YEAR - 10}-{_CURRENT_YEAR - 1}, Gaussian sampling)"
        )

    print(f"[Weather] Temperature data ready: {len(series):,} hourly records. Mode: {mode}")
    return series, mode, city_info


def resample_to_timestep(hourly_series: pd.Series, timestep_hours: float) -> np.ndarray:
    """
    Convert an hourly temperature Series to the requested timestep.
    - timestep_hours == 1   → return as-is (8,760 values)
    - timestep_hours == 0.5 → linear interpolation to 17,520 values
    """
    if timestep_hours == 1:
        return hourly_series.values

    if timestep_hours == 0.5:
        # Cubic interpolation at sub-hourly resolution
        n_hourly = len(hourly_series)
        x_hourly = np.arange(n_hourly, dtype=float)
        x_half   = np.arange(0, n_hourly - 0.5, 0.5)
        return np.interp(x_half, x_hourly, hourly_series.values)

    raise ValueError(f"timestep_hours must be 1 or 0.5, got {timestep_hours}")


# ═════════════════════════════════════════════════════════════════════════════
# DATA CENTRE MODEL  (physics unchanged from original paper)
# ═════════════════════════════════════════════════════════════════════════════

class DataCenterModel:
    """
    Physics-based data center power consumption model.

    Parameters
    ----------
    it_capacity_kw : float
        Maximum IT equipment capacity in kW.
    pue : float
        Target Power Usage Effectiveness (1.2-2.0 typical).
    base_utilization : float
        Baseline IT load fraction (0-1).
    """

    def __init__(self, it_capacity_kw=500, pue=1.4, base_utilization=0.65):
        self.it_capacity_kw    = it_capacity_kw
        self.pue               = pue
        self.base_utilization  = base_utilization
        self.total_capacity_kw = it_capacity_kw * pue
        self.infrastructure_fraction = 0.08   # 8 % of total capacity

    # ── IT load ───────────────────────────────────────────────────────────────

    def generate_it_load(self, n_steps: int, timestep_hours: float) -> np.ndarray:
        """Generate IT load profile with daily/weekly/random patterns (Eq. 2-7)."""
        time = np.arange(n_steps) * timestep_hours

        base   = self.base_utilization * np.ones(n_steps)
        daily  = 0.07 * np.sin(2 * np.pi * time / 24   - np.pi / 2)   # peak ~18:00
        weekly = 0.04 * np.sin(2 * np.pi * time / 168)                 # weekly pattern

        rng = np.random.default_rng(42)
        noise = rng.normal(0, 0.02, n_steps)

        utilization = np.clip(base + daily + weekly + noise, 0.40, 0.95)
        return utilization * self.it_capacity_kw

    # ── Cooling load ──────────────────────────────────────────────────────────

    def calculate_cooling_load(
        self, it_power_kw: np.ndarray, outdoor_temp_c: np.ndarray
    ) -> np.ndarray:
        """Temperature-dependent cooling load (Eq. 8-11)."""
        base_cooling = it_power_kw * (self.pue - 1) * 0.8          # Eq. 8
        temp_factor  = 1.0 + 0.02 * (outdoor_temp_c - 15.0)        # Eq. 9
        temp_factor  = np.clip(temp_factor, 0.7, 1.5)               # Eq. 10
        return base_cooling * temp_factor                            # Eq. 11

    # ── Full annual profile ───────────────────────────────────────────────────

    def generate_full_profile(
        self,
        outdoor_temp_c: np.ndarray,
        city_name: str,
        year: int,
        weather_mode: str,
        timestep_hours: float = 1,
    ) -> pd.DataFrame:
        """
        Assemble the complete data-centre power profile.

        Parameters
        ----------
        outdoor_temp_c : np.ndarray
            Outdoor temperature array already resampled to the target timestep.
        city_name      : str   - location label for the datetime index start
        year           : int   - simulation year
        weather_mode   : str   - description of weather data source
        timestep_hours : float - 1 or 0.5
        """
        n_steps = len(outdoor_temp_c)
        time    = np.arange(n_steps) * timestep_hours

        it_kw      = self.generate_it_load(n_steps, timestep_hours)
        cooling_kw = self.calculate_cooling_load(it_kw, outdoor_temp_c)
        infra_kw   = (
            self.total_capacity_kw * self.infrastructure_fraction * np.ones(n_steps)
        )
        total_kw   = it_kw + cooling_kw + infra_kw

        waste_heat_kw      = it_kw * 0.98 + cooling_kw * 0.10   # Eq. 16
        recoverable_kw     = waste_heat_kw * 0.40                # Eq. 17

        freq = f"{int(timestep_hours * 60)}min"
        dt_index = pd.date_range(
            start=f"{year}-01-01", periods=n_steps, freq=freq
        )

        df = pd.DataFrame(
            {
                "datetime":             dt_index,
                "time_h":               time,
                "it_load_kw":           it_kw,
                "cooling_load_kw":      cooling_kw,
                "infrastructure_kw":    infra_kw,
                "total_load_kw":        total_kw,
                "outdoor_temp_c":       outdoor_temp_c,
                "waste_heat_kw":        waste_heat_kw,
                "recoverable_heat_kw":  recoverable_kw,
                "actual_pue":           total_kw / it_kw,
            }
        )
        return df


# ═════════════════════════════════════════════════════════════════════════════
# PLOTTING
# ═════════════════════════════════════════════════════════════════════════════

def plot_profiles(df, city_info, year, weather_mode, timestep_hours, save_path=None):
    """Four-panel annual overview plot."""
    fig, axes = plt.subplots(4, 1, figsize=(14, 14))
    loc_label = f"{city_info['name']}, {city_info['country']} — {year}"
    fig.suptitle(
        f"Data Centre Energy Model | {loc_label}\n"
        f"Weather: {weather_mode}",
        fontsize=11, y=0.98
    )

    # 1. Annual power
    ax = axes[0]
    ax.plot(df["datetime"], df["total_load_kw"],   label="Total Power",   lw=0.6)
    ax.plot(df["datetime"], df["it_load_kw"],       label="IT Load",       lw=0.6, alpha=0.8)
    ax.plot(df["datetime"], df["cooling_load_kw"],  label="Cooling Load",  lw=0.6, alpha=0.8)
    ax.set_ylabel("Power (kW)")
    ax.set_title("Annual Power Consumption")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # 2. Typical week
    ax = axes[1]
    week_end = 168 / timestep_hours
    wk = df[df["time_h"] < week_end]
    ax.plot(wk["time_h"], wk["total_load_kw"],   label="Total Power",  lw=1.4)
    ax.plot(wk["time_h"], wk["it_load_kw"],       label="IT Load",      lw=1.4, alpha=0.8)
    ax.plot(wk["time_h"], wk["cooling_load_kw"],  label="Cooling Load", lw=1.4, alpha=0.8)
    ax.set_xlabel("Hours from start of year")
    ax.set_ylabel("Power (kW)")
    ax.set_title("Typical First-Week Profile")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    # 3. Temperature vs. cooling
    ax = axes[2]
    ax.plot(df["datetime"], df["outdoor_temp_c"], color="tomato", lw=0.6, label="Outdoor Temp")
    ax.set_ylabel("Temperature (°C)", color="tomato")
    ax.tick_params(axis="y", labelcolor="tomato")
    ax2 = ax.twinx()
    ax2.plot(df["datetime"], df["cooling_load_kw"], color="steelblue", lw=0.6, label="Cooling")
    ax2.set_ylabel("Cooling Power (kW)", color="steelblue")
    ax2.tick_params(axis="y", labelcolor="steelblue")
    ax.set_title("Outdoor Temperature vs. Cooling Load")
    ax.grid(True, alpha=0.3)

    # 4. Waste heat
    ax = axes[3]
    ax.plot(df["datetime"], df["waste_heat_kw"],     label="Total Waste Heat",  lw=0.6)
    ax.plot(df["datetime"], df["recoverable_heat_kw"], label="Recoverable Heat", lw=0.6)
    ax.set_xlabel("Date")
    ax.set_ylabel("Heat (kW)")
    ax.set_title("Waste Heat Generation (VHP Integration)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    plt.tight_layout(rect=[0, 0, 1, 0.96])

    if save_path:
        plt.savefig(save_path, dpi=200, bbox_inches="tight")
        print(f"[Output] Plot saved → {save_path}")
    plt.close(fig)
    return fig


# ═════════════════════════════════════════════════════════════════════════════
# SUMMARY PRINTER
# ═════════════════════════════════════════════════════════════════════════════

def print_summary(df, dc, city_info, year, weather_mode, timestep_hours):
    ts_label = "hourly" if timestep_hours == 1 else "half-hourly"
    print("\n" + "=" * 65)
    print("  DATA CENTRE ENERGY MODEL — ANNUAL SUMMARY")
    print("=" * 65)
    print(f"  Location     : {city_info['name']}, {city_info['country']}")
    print(f"  Year         : {year}")
    print(f"  Weather mode : {weather_mode}")
    print(f"  Timestep     : {ts_label} ({timestep_hours}h, {len(df):,} rows)")
    print("-" * 65)
    print(f"  IT Capacity      : {dc.it_capacity_kw:,.0f} kW")
    print(f"  Total Capacity   : {dc.total_capacity_kw:,.0f} kW  (PUE {dc.pue})")
    print("-" * 65)
    print("  Average IT Load      :", f"{df['it_load_kw'].mean():>8,.1f} kW",
          f"  ({df['it_load_kw'].mean() / dc.it_capacity_kw * 100:.1f}% utilisation)")
    print("  Average Cooling Load :", f"{df['cooling_load_kw'].mean():>8,.1f} kW")
    print("  Average Total Load   :", f"{df['total_load_kw'].mean():>8,.1f} kW")
    print("  Average PUE          :", f"{df['actual_pue'].mean():>8.3f}")
    print("  Min / Max PUE        :",
          f"{df['actual_pue'].min():.3f} / {df['actual_pue'].max():.3f}")
    print("-" * 65)
    print("  Avg Outdoor Temp     :", f"{df['outdoor_temp_c'].mean():>8.1f} °C")
    print("  Min / Max Temp       :",
          f"{df['outdoor_temp_c'].min():.1f} / {df['outdoor_temp_c'].max():.1f} °C")
    print("-" * 65)
    hrs_per_step = timestep_hours          # convert kW → kWh
    print("  Avg Waste Heat       :", f"{df['waste_heat_kw'].mean():>8,.1f} kW")
    print("  Avg Recoverable Heat :", f"{df['recoverable_heat_kw'].mean():>8,.1f} kW")
    annual_rec_mwh = df["recoverable_heat_kw"].sum() * hrs_per_step / 1000
    print("  Annual Recoverable   :", f"{annual_rec_mwh:>8,.1f} MWh")
    print("=" * 65)


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def save_excel(df, dc, city_info, year, weather_mode, timestep_hours, output_dir):
    """
    Save results to a formatted Excel workbook with three sheets:
      1. Metadata   - run configuration and summary statistics
      2. Hourly Data - full time-series profile
      3. Monthly Summary - monthly averages and totals
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    city_slug = city_info["name"].lower().replace(" ", "_").replace(",", "")
    ts_label  = "30min" if timestep_hours == 0.5 else "1h"
    path      = f"{output_dir}/dc_profile_{city_slug}_{year}_{ts_label}.xlsx"

    hrs = timestep_hours
    annual_rec_mwh = df["recoverable_heat_kw"].sum() * hrs / 1000

    # ── Sheet 1: Metadata ─────────────────────────────────────────────────────
    meta_rows = [
        ("Run Configuration", ""),
        ("City",              f"{city_info['name']}, {city_info['country']}"),
        ("Latitude",          f"{city_info['latitude']:.4f}"),
        ("Longitude",         f"{city_info['longitude']:.4f}"),
        ("Year",              year),
        ("Weather Mode",      weather_mode),
        ("Timestep",          f"{ts_label} ({timestep_hours} h)"),
        ("Total Records",     len(df)),
        ("", ""),
        ("Data Centre Parameters", ""),
        ("IT Capacity (kW)",       dc.it_capacity_kw),
        ("Target PUE",             dc.pue),
        ("Base Utilisation",       f"{dc.base_utilization:.0%}"),
        ("Total Capacity (kW)",    dc.total_capacity_kw),
        ("", ""),
        ("Annual Performance", ""),
        ("Avg IT Load (kW)",       round(df["it_load_kw"].mean(), 1)),
        ("Avg Cooling Load (kW)",  round(df["cooling_load_kw"].mean(), 1)),
        ("Avg Total Load (kW)",    round(df["total_load_kw"].mean(), 1)),
        ("Avg PUE",                round(df["actual_pue"].mean(), 3)),
        ("Min PUE",                round(df["actual_pue"].min(), 3)),
        ("Max PUE",                round(df["actual_pue"].max(), 3)),
        ("Avg Outdoor Temp (°C)",  round(df["outdoor_temp_c"].mean(), 1)),
        ("Avg Waste Heat (kW)",    round(df["waste_heat_kw"].mean(), 1)),
        ("Avg Recoverable Heat (kW)", round(df["recoverable_heat_kw"].mean(), 1)),
        ("Annual Recoverable Heat (MWh)", round(annual_rec_mwh, 1)),
    ]
    meta_df = pd.DataFrame(meta_rows, columns=["Parameter", "Value"])

    # ── Sheet 2: Monthly summary ───────────────────────────────────────────────
    df_m = df.copy()
    df_m["month"] = pd.to_datetime(df_m["datetime"]).dt.month
    monthly = df_m.groupby("month").agg(
        avg_it_kw        =("it_load_kw",           "mean"),
        avg_cooling_kw   =("cooling_load_kw",       "mean"),
        avg_total_kw     =("total_load_kw",         "mean"),
        avg_pue          =("actual_pue",            "mean"),
        avg_temp_c       =("outdoor_temp_c",        "mean"),
        avg_rec_heat_kw  =("recoverable_heat_kw",   "mean"),
        total_rec_mwh    =("recoverable_heat_kw",   lambda x: round(x.sum() * hrs / 1000, 1)),
    ).round(2).reset_index()
    monthly.insert(0, "month_name", pd.to_datetime(monthly["month"], format="%m").dt.strftime("%B"))
    monthly.columns = [
        "Month No.", "Month", "Avg IT (kW)", "Avg Cooling (kW)",
        "Avg Total (kW)", "Avg PUE", "Avg Temp (°C)",
        "Avg Rec. Heat (kW)", "Total Rec. Heat (MWh)",
    ]

    # ── Write workbook ─────────────────────────────────────────────────────────
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        meta_df.to_excel(writer, sheet_name="Metadata",        index=False)
        df.to_excel(     writer, sheet_name="Hourly Data",     index=False)
        monthly.to_excel(writer, sheet_name="Monthly Summary", index=False)

        wb = writer.book

        # ── Style helper ──────────────────────────────────────────────────────
        header_fill   = PatternFill("solid", fgColor="1F4E79")
        section_fill  = PatternFill("solid", fgColor="2E75B6")
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        section_font  = Font(bold=True, color="FFFFFF", size=10)
        normal_font   = Font(size=10)
        thin_border   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # ── Format Metadata sheet ─────────────────────────────────────────────
        ws = wb["Metadata"]
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 52
        for row in ws.iter_rows():
            for cell in row:
                cell.font   = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            # Section headers (col B empty, col A is a heading)
            a_val = str(row[0].value or "")
            if a_val in ("Run Configuration", "Data Centre Parameters", "Annual Performance"):
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_font
        # Top header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # ── Format Hourly Data sheet ──────────────────────────────────────────
        ws2 = wb["Hourly Data"]
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 22)
        ws2.freeze_panes = "A2"

        # ── Format Monthly Summary sheet ──────────────────────────────────────
        ws3 = wb["Monthly Summary"]
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws3.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws3.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 24)
        # Alternating row colours
        light_blue = PatternFill("solid", fgColor="DEEAF1")
        for i, row in enumerate(ws3.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = light_blue

    print(f"[Output] Excel saved → {path}")
    return path


def main(cfg: dict):
    city           = cfg["city"]
    year           = cfg["year"]
    timestep       = cfg["timestep_hours"]
    it_cap         = cfg["it_capacity_kw"]
    pue            = cfg["pue"]
    base_util      = cfg["base_utilization"]
    output_dir     = cfg["output_dir"]

    # ── 1. Fetch real weather data ────────────────────────────────────────────
    hourly_temp, weather_mode, city_info = fetch_temperature(city, year)

    # ── 2. Resample to requested timestep ─────────────────────────────────────
    temp_array = resample_to_timestep(hourly_temp, timestep)

    # ── 3. Build data-centre model ────────────────────────────────────────────
    dc = DataCenterModel(
        it_capacity_kw   = it_cap,
        pue              = pue,
        base_utilization = base_util,
    )

    # ── 4. Generate full annual profile ──────────────────────────────────────
    print("\n[Model] Generating data centre load profile ...")
    df = dc.generate_full_profile(
        outdoor_temp_c = temp_array,
        city_name      = city_info["name"],
        year           = year,
        weather_mode   = weather_mode,
        timestep_hours = timestep,
    )

    # ── 5. Print summary ──────────────────────────────────────────────────────
    print_summary(df, dc, city_info, year, weather_mode, timestep)

    # ── 6. Save CSV ───────────────────────────────────────────────────────────
    if cfg["output_csv"]:
        city_slug = city_info["name"].lower().replace(" ", "_").replace(",", "")
        ts_label  = "30min" if timestep == 0.5 else "1h"
        csv_path  = f"{output_dir}/dc_profile_{city_slug}_{year}_{ts_label}.csv"
        df.to_csv(csv_path, index=False)
        print(f"[Output] CSV saved  → {csv_path}")

    # ── 7. Save Excel ─────────────────────────────────────────────────────────
    if cfg.get("output_excel"):
        save_excel(df, dc, city_info, year, weather_mode, timestep, output_dir)

    # ── 8. Save plots ─────────────────────────────────────────────────────────
    if cfg["output_plots"]:
        city_slug  = city_info["name"].lower().replace(" ", "_").replace(",", "")
        ts_label   = "30min" if timestep == 0.5 else "1h"
        plot_path  = f"{output_dir}/dc_profiles_{city_slug}_{year}_{ts_label}.png"
        plot_profiles(df, city_info, year, weather_mode, timestep, save_path=plot_path)

    return df, dc, city_info, weather_mode


if __name__ == "__main__":
    df, dc, city_info, weather_mode = main(CONFIG)
